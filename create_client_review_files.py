"""
Create Client Review Files
- Data quality comparison (before/after normalization)
- Excel file with highlighted rows for client review
- Updated data quality summary
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

print("="*70)
print("CREATING CLIENT REVIEW FILES")
print("="*70)

# =============================================================================
# LOAD DATA
# =============================================================================
print("\nLoading data...")

# Original data (case-level)
df_orig = pd.read_csv('prepared_data/bc_national_comparison.csv')
df_orig_merged = pd.read_csv('prepared_data/pricing_sales_merged.csv')

# Normalized data (unit-level)
df_norm = pd.read_csv('prepared_data_normalized/bc_national_comparison.csv')
df_norm_merged = pd.read_csv('prepared_data_normalized/pricing_sales_merged.csv')

# Product master
df_products = pd.read_csv('prepared_data_normalized/product_master.csv')

print(f"  Original BC-National: {len(df_orig):,} rows")
print(f"  Normalized BC-National: {len(df_norm):,} rows")

# =============================================================================
# IDENTIFY ISSUES THAT WERE RESOLVED BY NORMALIZATION
# =============================================================================
print("\nAnalyzing data quality changes...")

# Issues in original data
df_orig_valid = df_orig[df_orig['price_gap_pct'].notna()].copy()

# Original issues
orig_negative = df_orig_valid[df_orig_valid['price_gap_pct'] < 0].copy()
orig_extreme_high = df_orig_valid[df_orig_valid['price_gap_pct'] > 60].copy()
orig_extreme_low = df_orig_valid[df_orig_valid['price_gap_pct'] < -100].copy()

print(f"\nORIGINAL DATA (Case-Level BSP):")
print(f"  Negative gaps: {len(orig_negative):,}")
print(f"  Extreme high gaps (>60%): {len(orig_extreme_high):,}")
print(f"  Extreme low gaps (<-100%): {len(orig_extreme_low):,}")

# Issues in normalized data
df_norm_valid = df_norm[df_norm['price_gap_pct'].notna()].copy()

norm_negative = df_norm_valid[df_norm_valid['price_gap_pct'] < 0].copy()
norm_extreme_high = df_norm_valid[df_norm_valid['price_gap_pct'] > 60].copy()
norm_extreme_low = df_norm_valid[df_norm_valid['price_gap_pct'] < -100].copy()

print(f"\nNORMALIZED DATA (Unit-Level BSP = BSP/Pack):")
print(f"  Negative gaps: {len(norm_negative):,}")
print(f"  Extreme high gaps (>60%): {len(norm_extreme_high):,}")
print(f"  Extreme low gaps (<-100%): {len(norm_extreme_low):,}")

# =============================================================================
# IDENTIFY RESOLVED ISSUES (were issues before, not issues now)
# =============================================================================
print("\nIdentifying resolved issues...")

# Create unique keys for comparison
df_orig_valid['key'] = df_orig_valid['item_code'].astype(str) + '_' + df_orig_valid['division'] + '_' + df_orig_valid['quarter']
df_norm_valid['key'] = df_norm_valid['item_code'].astype(str) + '_' + df_norm_valid['division'] + '_' + df_norm_valid['quarter']

# Also add keys to the issue subsets
orig_negative['key'] = orig_negative['item_code'].astype(str) + '_' + orig_negative['division'] + '_' + orig_negative['quarter']
norm_negative['key'] = norm_negative['item_code'].astype(str) + '_' + norm_negative['division'] + '_' + norm_negative['quarter']

# Find rows that were negative before but not after
orig_negative_keys = set(orig_negative['key'])
norm_negative_keys = set(norm_negative['key'])
resolved_negative_keys = orig_negative_keys - norm_negative_keys

print(f"\nRESOLVED ISSUES (fixed by normalization):")
print(f"  Negative gaps resolved: {len(resolved_negative_keys):,}")

# Get the resolved rows from original data
resolved_rows = df_orig_valid[df_orig_valid['key'].isin(resolved_negative_keys)].copy()

# Add normalized values for comparison (drop duplicates to avoid index error)
norm_lookup = df_norm_valid.drop_duplicates(subset='key').set_index('key')[['price_gap_pct', 'bsp_per_unit', 'national_bsp_per_unit', 'pack', 'national_pack']].to_dict('index')

resolved_rows['normalized_gap_pct'] = resolved_rows['key'].map(lambda k: norm_lookup.get(k, {}).get('price_gap_pct', np.nan))
resolved_rows['bsp_per_unit'] = resolved_rows['key'].map(lambda k: norm_lookup.get(k, {}).get('bsp_per_unit', np.nan))
resolved_rows['national_bsp_per_unit'] = resolved_rows['key'].map(lambda k: norm_lookup.get(k, {}).get('national_bsp_per_unit', np.nan))
resolved_rows['bc_pack'] = resolved_rows['key'].map(lambda k: norm_lookup.get(k, {}).get('pack', np.nan))
resolved_rows['national_pack'] = resolved_rows['key'].map(lambda k: norm_lookup.get(k, {}).get('national_pack', np.nan))

resolved_rows['issue_type'] = 'Negative Gap - RESOLVED by unit normalization'
resolved_rows['resolution'] = 'Pack size difference explains the gap'

print(f"  Rows prepared for client review: {len(resolved_rows):,}")

# =============================================================================
# IDENTIFY REMAINING ISSUES (still issues after normalization)
# =============================================================================
print("\nIdentifying remaining issues...")

remaining_issues = df_norm_valid[
    (df_norm_valid['price_gap_pct'] < 0) |
    (df_norm_valid['price_gap_pct'] > 80)
].copy()

remaining_issues['issue_type'] = np.where(
    remaining_issues['price_gap_pct'] < 0,
    'Negative Gap - REMAINS after normalization',
    'Extreme High Gap (>80%) - REMAINS after normalization'
)

print(f"  Remaining issues: {len(remaining_issues):,}")

# =============================================================================
# CREATE EXCEL FILE WITH HIGHLIGHTING
# =============================================================================
print("\nCreating Excel file for client review...")

wb = Workbook()

# Define styles
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# --- Sheet 1: Summary ---
ws_summary = wb.active
ws_summary.title = "Summary"

summary_data = [
    ["AWG Data Quality Analysis - Client Review", ""],
    ["", ""],
    ["ISSUE CATEGORY", "COUNT"],
    ["", ""],
    ["ORIGINAL DATA (Case-Level BSP)", ""],
    ["Negative price gaps (BC > National)", len(orig_negative)],
    ["Extreme gaps (>60%)", len(orig_extreme_high)],
    ["", ""],
    ["AFTER NORMALIZATION (Unit-Level BSP = BSP/Pack)", ""],
    ["Negative price gaps", len(norm_negative)],
    ["Extreme gaps (>60%)", len(norm_extreme_high)],
    ["", ""],
    ["RESOLVED BY NORMALIZATION", ""],
    ["Issues that went away after using BSP/Pack", len(resolved_negative_keys)],
    ["", ""],
    ["REMAINING ISSUES (Need Client Review)", ""],
    ["Negative gaps still present", len(norm_negative)],
    ["", ""],
    ["KEY FINDING", ""],
    ["The original 'issues' were mostly due to comparing CASE prices", ""],
    ["(BSP) vs UNIT prices (SRP, National BSP for different pack sizes)", ""],
    ["", ""],
    ["RECOMMENDATION", ""],
    ["Normalize all prices to unit level using Pack (Units per Case)", ""],
]

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif "ORIGINAL" in str(value) or "AFTER" in str(value) or "RESOLVED" in str(value) or "REMAINING" in str(value) or "KEY FINDING" in str(value) or "RECOMMENDATION" in str(value):
            cell.font = Font(bold=True)

ws_summary.column_dimensions['A'].width = 60
ws_summary.column_dimensions['B'].width = 15

# --- Sheet 2: Resolved Issues (Green highlight - fixed by normalization) ---
ws_resolved = wb.create_sheet("Resolved_Issues_GREEN")

# Select columns for resolved issues
resolved_cols = ['item_code', 'item_name', 'category', 'division', 'quarter',
                 'bsp', 'national_bsp', 'price_gap_pct',
                 'bc_pack', 'national_pack',
                 'bsp_per_unit', 'national_bsp_per_unit', 'normalized_gap_pct',
                 'issue_type', 'resolution']

resolved_export = resolved_rows[[c for c in resolved_cols if c in resolved_rows.columns]].copy()
resolved_export = resolved_export.rename(columns={
    'bsp': 'BC_BSP_Case',
    'national_bsp': 'National_BSP_Case',
    'price_gap_pct': 'Original_Gap_Pct',
    'bc_pack': 'BC_Pack_Size',
    'national_pack': 'National_Pack_Size',
    'bsp_per_unit': 'BC_BSP_Unit',
    'national_bsp_per_unit': 'National_BSP_Unit',
    'normalized_gap_pct': 'Normalized_Gap_Pct'
})

# Write headers
headers = list(resolved_export.columns)
for col_idx, header in enumerate(headers, 1):
    cell = ws_resolved.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

# Write data with green highlighting
for row_idx, row in enumerate(resolved_export.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        cell = ws_resolved.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = green_fill
        cell.border = thin_border

# Adjust column widths
for col_idx, header in enumerate(headers, 1):
    ws_resolved.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = max(len(str(header)) + 2, 12)

# --- Sheet 3: Remaining Issues (Red highlight - need client review) ---
ws_remaining = wb.create_sheet("Remaining_Issues_RED")

remaining_cols = ['item_code', 'item_name', 'category', 'division', 'quarter',
                  'bsp', 'national_bsp', 'bsp_per_unit', 'national_bsp_per_unit',
                  'price_gap_pct', 'pack', 'issue_type']

remaining_export = remaining_issues[[c for c in remaining_cols if c in remaining_issues.columns]].copy()
remaining_export = remaining_export.rename(columns={
    'bsp': 'BC_BSP_Case',
    'national_bsp': 'National_BSP_Case',
    'bsp_per_unit': 'BC_BSP_Unit',
    'national_bsp_per_unit': 'National_BSP_Unit',
    'price_gap_pct': 'Gap_Pct_Unit_Level',
    'pack': 'BC_Pack_Size'
})

# Write headers
headers = list(remaining_export.columns)
for col_idx, header in enumerate(headers, 1):
    cell = ws_remaining.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

# Write data with red highlighting
for row_idx, row in enumerate(remaining_export.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        cell = ws_remaining.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = red_fill
        cell.border = thin_border

# Adjust column widths
for col_idx, header in enumerate(headers, 1):
    col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
    ws_remaining.column_dimensions[col_letter].width = max(len(str(header)) + 2, 12)

# --- Sheet 4: Comparison Statistics ---
ws_stats = wb.create_sheet("Statistics_Comparison")

stats_data = [
    ["Metric", "Original (Case BSP)", "Normalized (Unit BSP)", "Change"],
    ["", "", "", ""],
    ["Average Price Gap", f"{df_orig_valid['price_gap_pct'].mean():.1f}%", f"{df_norm_valid['price_gap_pct'].mean():.1f}%", f"{df_norm_valid['price_gap_pct'].mean() - df_orig_valid['price_gap_pct'].mean():+.1f}%"],
    ["Median Price Gap", f"{df_orig_valid['price_gap_pct'].median():.1f}%", f"{df_norm_valid['price_gap_pct'].median():.1f}%", f"{df_norm_valid['price_gap_pct'].median() - df_orig_valid['price_gap_pct'].median():+.1f}%"],
    ["Std Dev", f"{df_orig_valid['price_gap_pct'].std():.1f}%", f"{df_norm_valid['price_gap_pct'].std():.1f}%", f"{df_norm_valid['price_gap_pct'].std() - df_orig_valid['price_gap_pct'].std():+.1f}%"],
    ["", "", "", ""],
    ["Negative Gaps (Count)", len(orig_negative), len(norm_negative), f"{len(norm_negative) - len(orig_negative):+,}"],
    ["Negative Gaps (%)", f"{100*len(orig_negative)/len(df_orig_valid):.1f}%", f"{100*len(norm_negative)/len(df_norm_valid):.1f}%", ""],
    ["", "", "", ""],
    ["Issues Resolved", "", len(resolved_negative_keys), ""],
    ["Issues Remaining", "", len(remaining_issues), ""],
]

for row_idx, row_data in enumerate(stats_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
        cell.border = thin_border

for col in ['A', 'B', 'C', 'D']:
    ws_stats.column_dimensions[col].width = 25

# Save workbook
output_file = 'Client_Review_Data_Quality_Issues.xlsx'
wb.save(output_file)
print(f"\nSaved: {output_file}")

# =============================================================================
# CREATE UPDATED DATA QUALITY CSV
# =============================================================================
print("\nCreating updated data quality CSV...")

# Combine resolved and remaining issues
all_issues = []

# Add resolved issues
for _, row in resolved_rows.iterrows():
    all_issues.append({
        'status': 'RESOLVED',
        'issue_type': row['issue_type'],
        'item_code': row['item_code'],
        'item_name': row.get('item_name', ''),
        'category': row.get('category', ''),
        'division': row['division'],
        'quarter': row['quarter'],
        'original_gap_pct': row['price_gap_pct'],
        'normalized_gap_pct': row['normalized_gap_pct'],
        'bc_bsp_case': row['bsp'],
        'national_bsp_case': row['national_bsp'],
        'bc_pack': row['bc_pack'],
        'national_pack': row['national_pack'],
        'bc_bsp_unit': row['bsp_per_unit'],
        'national_bsp_unit': row['national_bsp_per_unit'],
        'resolution': 'Pack size normalization'
    })

# Add remaining issues
for _, row in remaining_issues.iterrows():
    all_issues.append({
        'status': 'REMAINING',
        'issue_type': row['issue_type'],
        'item_code': row['item_code'],
        'item_name': row.get('item_name', ''),
        'category': row.get('category', ''),
        'division': row['division'],
        'quarter': row['quarter'],
        'original_gap_pct': np.nan,
        'normalized_gap_pct': row['price_gap_pct'],
        'bc_bsp_case': row['bsp'],
        'national_bsp_case': row['national_bsp'],
        'bc_pack': row.get('pack', np.nan),
        'national_pack': row.get('national_pack', np.nan),
        'bc_bsp_unit': row.get('bsp_per_unit', np.nan),
        'national_bsp_unit': row.get('national_bsp_per_unit', np.nan),
        'resolution': 'Needs client review'
    })

df_all_issues = pd.DataFrame(all_issues)
df_all_issues.to_csv('prepared_data_normalized/Data_Quality_Issues_Updated.csv', index=False)
print(f"Saved: prepared_data_normalized/Data_Quality_Issues_Updated.csv ({len(df_all_issues):,} rows)")

# =============================================================================
# SUMMARY
# =============================================================================
print("\n" + "="*70)
print("FILES CREATED")
print("="*70)
print(f"""
1. Client_Review_Data_Quality_Issues.xlsx
   - Summary tab: Overview of issues
   - Resolved_Issues_GREEN: {len(resolved_rows):,} rows (green highlight)
   - Remaining_Issues_RED: {len(remaining_issues):,} rows (red highlight)
   - Statistics_Comparison: Before/after metrics

2. prepared_data_normalized/Data_Quality_Issues_Updated.csv
   - Combined resolved + remaining issues
   - {len(df_all_issues):,} total rows
""")

print("="*70)
print("READY FOR CLIENT REVIEW")
print("="*70)
